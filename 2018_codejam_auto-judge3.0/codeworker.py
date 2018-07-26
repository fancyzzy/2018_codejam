#!/usr/bin/env python
# -*- coding=utf-8 -*-

import subprocess
import os
import time
import re
from threading import Timer
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from six import string_types

Run_In_Detail = False


def addTestCase(caseDir, case_list):
    '''
    Add all the test cases
    '''
    #print("DEBUG caseDir: {}".format(caseDir))
    os.chdir(caseDir)
    filelist = os.listdir('.')
    file_list = []
    for i in range(len(filelist)):
        num = re.findall(r'\d+',filelist[i])
        if len(num) != 0:
            file_list.append((filelist[i],int(num[0])))

    #sort according to the number in the Case name
    file_list = sorted(file_list, key=lambda x:x[1])
    filelist = [row[0] for row in file_list]

    for aFile in filelist:
        caseFile = open(aFile)
        try: # read all files into caseTxt
            testParam = ''
            testRes = ''
            testScore = '0'
            for line in caseFile.readlines():
                if line in ['\r\n', '\r', '\n']: #empty line
                    continue
                elif line.find("#SCORE") != -1:
                    testScore += line.replace('#SCORE:', '').strip('\n')
                elif line.find("#RES") != -1:
                    testRes = line
                else:
                    testParam = testParam + line
            testRes2=testRes.replace('#RES:', '').strip('\n')

            #case_list = [(case_name,input,output,score)]
            case_list.append((aFile,testParam,testRes2,int(testScore)))

        finally:
            caseFile.close()
    #print("DEBUG case_list :{}".format(case_list))

    #Use excel to add test cases:    
    #os.chdir(caseDir.strip('case'))
    test_case_list = []
    case_file_name = 'testcases.xlsx'
    case_file = os.path.join(caseDir.strip('case'),case_file_name)
    #print("DEBUG case_file: {}".format(case_file))

    wb = None
    try:
        wb = openpyxl.load_workbook(case_file, data_only=True)
    except:
        return []
    sheets = wb.sheetnames[:]
    #find the case sheet and the col, row
    fields = ['Index', 'Input', 'Expected', 'Score']
    row_range = 10
    column_range = 10
    base_row = None
    base_col = None
    base_ws = None
    have_found = False
    for sheet in sheets:
        for i in range(1, row_range):
            for j in range(1, column_range):
                v = wb[sheet].cell(row=i, column=j).value
                v_next = wb[sheet].cell(row=i, column=j+1).value
                if isinstance(v, string_types) and isinstance(v_next, string_types):
                    if  'input' in v.lower() and 'expect' in v_next.lower(): 
                        have_found = True
                        base_row = i
                        base_col = j
                        base_ws = wb[sheet]

    if not have_found:
        print("No Test cases info found (search column fields like: input, expect)!")
        return []
    else:

        case_max = 20
        n = 0
        for i in range(1, case_max+1):
            input_v = base_ws.cell(row=base_row+i, column=base_col).value
            expected_v = base_ws.cell(row=base_row+i, column=base_col+1).value
            if input_v == None and expected_v == None:
                break
            else:
                input_v = str(input_v)
                expected_v = str(expected_v)
                score_v = base_ws.cell(row=base_row+i, column=base_col+2).value
                if score_v != None and isinstance(score_v, string_types) and\
                score_v.isdigit():
                    score_v = int(score_v)
            test_case_list.append(('case%d'%i, input_v, expected_v, score_v))


        #print("DEBUG test_case_list = {}".format(test_case_list))
    return test_case_list


##########add_test_cases()##############################


def addExe(exeDir):
    '''
    List all the to be tested files
    '''

    ignore_suffix = ['.xlsx', '.docx', '.pptx', '.db']

    return [f for f in os.listdir(exeDir) if (not f.startswith('.') and  \
    os.path.splitext(f)[1] not in ignore_suffix) and os.path.isfile(os.path.join(exeDir,f))]


#feature: terminalt after timeout begin
def timer_expiry(proc,f):

    #print "timer expired"
    f.append('expiry')

    ppid = str(proc.pid)
    cmd = ['tskill',ppid]
    #Windows terminate a process using 'tskill', type 'tasklist' in cmd to see processes
    p = subprocess.Popen(cmd,stdout=subprocess.PIPE,stderr=subprocess.PIPE)

    err = p.communicate()[1]

    #print "DEBUG pid {} was killed".format(proc.pid)    

    if err == '':
        #print "%s was terminated."%(str(ppid))
        pass
    else:
        print "terminate process failed, err=",err

#feature: terminalt after timeout end



def callProc(cmd, inStr, outStr, trial_run = False):
    '''
    input: 
    cmd: trigger cmd
    inStr: input string
    outStr: Expected output string
    trial_run: to judge if print log

    '''
    global Run_In_Detail

    result = 1

    #print "debug cmd:",cmd
    p1 = subprocess.Popen(cmd, stdin=subprocess.PIPE, stdout=subprocess.PIPE,\
     stderr=subprocess.PIPE, shell=False)

    #feature: terminalt after timeout begin
    f = []
    my_timer = Timer(5,timer_expiry,[p1,f])
    try:
        my_timer.start()
        realOutStr, errStr = p1.communicate(input=inStr)
    except:
        print "Debug:catch except"
    finally:
        my_timer.cancel()    
    #feature: terminalt after timeout end

    #get rid of '\n\r' in windows
    realOutStr = realOutStr.strip(os.linesep)

    #Right answers are inferior to errors
    if errStr != '':
        result = 0
        return result,realOutStr,errStr

    if cmp(outStr, realOutStr)!=0:
        '''
        if Run_In_Detail and not trial_run:
            print("\nYour output: {}".format(realOutStr))
            print("   Expected: {}".format(outStr))
        '''
        if outStr in realOutStr:
            return result,realOutStr,errStr
        else:
            result = 0
    else:
        return result,realOutStr,errStr


    #timer expired accounted as TLE
    if len(f) == 1 and f[0] == 'expiry':
        result = 0
        errStr = "5s timer expired"

    return result,realOutStr,errStr


def execTest():
    timeMsLimit = 1500    # 1000ms = 1S
    #These dir can be used only after the system environment configured.
    #Check out the readme to see how to set system environment variables.
    python2Dir = 'python2'
    python3Dir = 'python3'
    javaDir = 'java'
    javacDir = 'javac'
    gccDir = 'gcc'
    gppDir = 'g++'    
    #Felix change the path begin    
    exeDir      = os.path.join(os.getcwd(),'exe')
    testCaseDir = os.path.join(os.getcwd(),'case')
    #Felix change the path end

    #Get all the to-be-test sourcecode/exe files
    exe_file_list = addExe(exeDir)
    #print("DEBUGF exe_file_list: {}".format(exe_file_list))
    print("Got {} to-be-test files from '{}'".format(len(exe_file_list), exeDir))

    case_list = []

    #case_list = [(case_name,input,output,score)]

    test_case_list = addTestCase(testCaseDir, case_list)
    if (not test_case_list and not case_list) or (len(test_case_list) == 0 and len(case_list) == 0):
        print("No test cases.")
        exit()

    if len(test_case_list) != 0:
        case_list = test_case_list
        case_file_name = 'testcases.xlsx'
        case_file = os.path.join(testCaseDir.strip('case'),case_file_name)
        testCaseDir = case_file


    case_total = len(case_list)    
    file_total = len(exe_file_list)
    score_total = sum([item[-1] for item in case_list])

    print("Prepared {} test cases, {} total points from '{}'".\
        format(case_total, score_total, testCaseDir))


    #case_result_list =[(code_name,language,score,run_time,state)]
    case_result_list = []
    submit_total = 0

    print("Start processing...")
    os.chdir(exeDir)
    # beging to loop everyone's code submission
    for exe_file in exe_file_list:

        file_exe = exe_file
        case_result = []
        score = 0
        time_used = []
        case_passed = 0
        case_error = 0
        case_tle = 0
        case_wrong = 0
        language =  'N/A'
        compile_success = True
        state = ''

        cmd = []

        a_detail_record = []

        if '.exe' in exe_file and ((exe_file[:-4] + '.cpp') in exe_file_list or \
                (exe_file[:-4] + '.c' in exe_file_list)):
                #exclude .exe file that was built previously
                continue

        elif '.class' in exe_file and ((exe_file[:-6] + '.java') in exe_file_list):
                #exclude .class file that was built previously
                continue

        submit_total += 1
        print "---"*20
        print "{}.'{}'".format(submit_total,exe_file)

        if '.py' in exe_file:
            #pythnon2,3 auto running: py -2 xxxpy2.py or py -3 xxxpy3.py
            if 'py3' in exe_file or '3.' in exe_file \
            or '3' in exe_file.replace('32','').replace('q3','').replace('Q3','').replace('.3',''):
                cmd = [python3Dir,exe_file]
                language = "Python3"
            else:
                cmd = [python2Dir,exe_file]
                language = "Python2"

        elif '.class' in exe_file: # java: java -cp . xxx(.class)
            cmd = [javaDir,'-cp','.',exe_file.strip('.class')]
            language = "Java"

        elif '.exe' in exe_file:
            cmd = [exe_file]
            language = "C/C++"

        #feature auto-compile C/C++/Java begin
        elif '.c' == exe_file[-2:]:
            language = "C"
            #'gcc -o xxx_c xxx.c' to build out a xxx_c.exe
            compile_cmd = [gccDir,'-std=c11','-o',exe_file.replace('.c',''),exe_file]
            p_compile = subprocess.Popen(compile_cmd, stdin=subprocess.PIPE, stdout=subprocess.PIPE,\
             stderr=subprocess.PIPE, shell=False)    

            p_compile.wait()
            err = p_compile.stderr.read()
            if 'error' not in err:
                err = ''

            if err == '':
                file_exe = exe_file.replace('.c','.exe')
                print "Successfully Compiled:'gcc -std=c11 -o {} {}'".format(file_exe.replace('.exe',''),exe_file)
                cmd = [file_exe]
            else:
                print "Compiled failed, error:"
                print err
                compile_success = False
                state = "Compile_Error"

        elif '.cpp' == exe_file[-4:]:
            language = "C++"
            #feature compile C/C++/Java begin
            #gcc -o xxx_c xxx.c to build out a xxx_c.exe
            compile_cmd = [gppDir,'-std=c++11','-o',exe_file.replace('.cpp',''),exe_file]
            p_compile = subprocess.Popen(compile_cmd, stdin=subprocess.PIPE, stdout=subprocess.PIPE,\
            stderr=subprocess.PIPE, shell=False)    
            p_compile.wait()

            err = p_compile.stderr.read()
            if 'error' not in err:
                err = ''
            if err == '':
                file_exe = exe_file.replace('.cpp','.exe')
                print "Successfully Compiled:'g++ -std=c++11 -o {} {}'".format(file_exe.replace('.exe',''),exe_file)
                cmd = [file_exe]
            else:
                print "Compiled failed, error:"
                print err
                compile_success = False

        elif '.java' in exe_file:
            language = "Java"
            #javac xxx.java build out a xxx.class
            compile_cmd = [javacDir,exe_file]
            p_compile = subprocess.Popen(compile_cmd, stdin=subprocess.PIPE, stdout=subprocess.PIPE,\
            stderr=subprocess.PIPE, shell=True)    
            p_compile.wait()

            err = p_compile.stderr.read()
            if 'error' not in err:
                err = ''
            if err == '':
                file_exe = exe_file.replace('.java','.class')
                print "Successfully compiled '{}' to '{}'".format(exe_file, file_exe)
                #cmd = 'java' + ' -cp . ' + file_exe.strip('.class')
                cmd = ['java','-cp','.',file_exe.strip('.class')]
            else:
                print "Compiled failed, error:"
                print err
                compile_success = False
         #feature compile C/C++/Java end

        elif '.sh' in exe_file:
            #need to vi and :set fileformat=unix or else there is /r error
            cmd = ['bash',exe_file]
            language = "bash"


        else:
            print "error, {0} is not supported to run.".format(exe_file)
            language = exe_file.split('.')[1]
            case_result.append("'{}'".format(exe_file))
            case_result.append(language)
            case_result.append('0')
            case_result.append('N/A')
            case_result.append('Unsupport Error')
            case_result_list.append(case_result)

            continue

        case_result.append("{}".format(exe_file))
        case_result.append(language)

        if not compile_success:
            print ""
            case_result.append(str(score))
            case_result.append('N/A')
            case_result.append('Compile_Error')
            case_result_list.append(case_result)
            print ""
            continue

        #Just run case1 to load code into memory    
        for case in case_list:
            s_input = case[1]
            s_output = case[2]
            callProc(cmd, s_input, s_output, trial_run=True)
            break

        #felix beging to loop to check every cases in case_list
        wrong_case = []
        tle_case = []
        error_case = []
        pass_case = []
        i = 0
        for case in case_list:
            i += 1

            case_name = case[0]
            s_input = case[1]
            s_output = case[2]
            case_score = case[3]

            s = "'{0}'({1} points): ".format(case_name, case_score)
            print "%-30s"%s,
            start_time = time.clock()

            re,real_output,err = callProc(cmd, s_input, s_output)

            cost_time = int((time.clock() - start_time)*1000)

            #re != 1, the right result is prior to errors
            if err != '' and re != 1:
                print "Run Error:",
                print err
                case_error += 1
                error_case.append(i)

            elif cost_time > timeMsLimit and re == 1:
                print "Time limit Exceeded: %dms"%(cost_time)
                time_used.append(cost_time)
                case_tle += 1
                tle_case.append(i)

            elif re == 0 and err == '':
                print "Wrong anwser. {}ms".format(cost_time)
                time_used.append(cost_time)
                case_wrong += 1
                wrong_case.append(i)
                if Run_In_Detail:
                    print("Your output:\n{}".format(real_output))
                    print("Expected:\n{}".format(s_output))

            elif re == 1 and cost_time < timeMsLimit:
                case_passed += 1
                score += case_score
                pass_case.append(i)
                print "'Passed!' {0}ms".format(cost_time)
                time_used.append(cost_time)
            else:
                print "'DEBUG:It could not happen'"

        if len(time_used) == 0:
            run_time = 'N/A'
        else:
            run_time = sum(time_used)/len(time_used)

        print "'{}' Result: {}/{} passed, {}/{} points, average {}ms"\
        .format(exe_file,case_passed,case_total,score,score_total,run_time)

        case_result.append(str(score))
        case_result.append(str(run_time))
        if case_passed > 0 and case_passed < case_total:
            state += '{} Passed:({}), '.format(case_passed, ', '.join(map(str, pass_case)))
        if case_tle > 0:
            state += '{} TLE:({}), '.format(case_tle, ', '.join(map(str, tle_case)))
        if case_wrong > 0:
            state += '{} Wrong:({}), '.format(case_wrong, ', '.join(map(str, wrong_case)))
        if case_error > 0:
            state += '{} Error:({}), '.format(case_error, ', '.join(map(str, error_case)))
        if case_passed == case_total:
            state = "All {} cases passed!".format(case_total)
        case_result.append(state.strip(', '))
        case_result_list.append(case_result)
        print ""
    #end for for exe_file in exe_file_list:

    print "Summary"
    print "-"*90
    print "%-5s"%("Index"),
    print "%-35s"%("File_Name"),
    print "%-10s"%("Language"),
    print "%-10s"%("Score"),
    print "%-10s"%("Run_Time"),
    print "%-10s"%("State")

    pass_num = 0
    index = 0
    for item in case_result_list:
        if 'All' in item[4]:
            pass_num += 1
        index += 1
        print "%-5s"%(str(index)),
        #print "{}.".format(j),
        for i in range(len(item)):
            if i == 0:
                print "%-35s"%(item[i]),
            else:
                print "%-10s"%(item[i]),
        print
    print "-"*90

    print "{} submissions checked, {} totally passed.".format(submit_total,pass_num)
    return case_result_list

########################execTest()######################################


def gen_report(res_list):
    '''
    Generate the result report in xlsx
    '''
    report_name = "result.xlsx"
    report_path = os.path.join(os.getcwd(), report_name)

    if  not os.path.exists(report_name):
        wb = openpyxl.Workbook(write_only=False)
        wb.title = report_name
    else:
        wb = openpyxl.load_workbook(report_name, data_only=True)

    #Generate time as the sheet name:
    sheet_name = time.strftime("%Y%m%d_%H%M%S")
    ws_new = wb.create_sheet(sheet_name, 0)

    #get results
    fields = ['Index', 'File_Name', 'Language', 'Score', 'Run_Time', 'State']
    ln_fields = len(fields)
    ln_res = len(res_list)
    #write results
    start_row = 2
    start_col = 1
    redFill = PatternFill(start_color='92D050', end_color='92D050',\
                   fill_type='solid')
    for j in range(ln_fields):
        #ws_new.cell(column=start_col + j, row=start_row, value=fields[j])
        ws_new.cell(row=start_row, column=start_col+j).value = fields[j]
        ws_new.cell(row=start_row, column=start_col+j).fill = redFill


    for i in range(ln_res):
        one_res = res_list[i]
        for j in range(ln_fields-1):
            ws_new.cell(column=start_col, row=start_row + 1 + i, value=i+1)
            ws_new.cell(column=start_col + 1 + j, row=start_row + 1 + i, \
                value=one_res[j])

    #set column width
    for j in range(ln_fields):
        width_value = 10
        if j == 1 or j == 5:
            width_value = 50
        else:
            pass
        i = get_column_letter(j+1)
        ws_new.column_dimensions[i].width = width_value

    #save after the file is closed    
    wb.save(report_name)

    print("Report is generated in '{}'".format(report_path))
##########gen_report()#################################


if __name__ == '__main__':

    n = raw_input("Run in details? (y/n): ")
    if n == 'y' or n == 'Y':
        Run_In_Detail = True
    else:
        Run_In_Detail = False

    res_list = execTest()

    n = raw_input("\nDo you want to generate/update the report? (y/n): \n(Close it if alreay opened by the Excel app)")
    if 'y' in n.strip() or n.strip() == 'Y':
        gen_report(res_list)
    else:
        pass


    print("Done")
